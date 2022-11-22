import time
from datetime import datetime

import requests
from openpyxl.styles import Font
from openpyxl.styles import PatternFill

whiteFill = PatternFill(fgColor="0070C0", bgColor="0070C0", fill_type="solid")
orangeFill = PatternFill(fgColor="FF462B", bgColor="FF462B", fill_type="solid")
greyFill = PatternFill(fgColor="D9D9D9", bgColor="D9D9D9", fill_type="solid")
grey_fill_light= PatternFill(fgColor="F2F2F2", bgColor="F2F2F2", fill_type="solid")
text_color = Font(color="FFFFFF")
text_color_red = Font(color="FF0000")
yellowFill = PatternFill(fgColor="FFFF00", bgColor="FFFF00", fill_type="solid")    
yellow_fill_light = PatternFill(fgColor="FFFF99", bgColor="FFFF99", fill_type="solid")
grey_date = PatternFill(fgColor="D9D9D9", bgColor="D9D9D9", fill_type="solid")
text_bold = Font(bold=True)# mettre la police en gras

# Documentation: https://developers.google.com/speed/docs/insights/v5/get-started

# JSON paths: https://developers.google.com/speed/docs/insights/v4/reference/pagespeedapi/runpagespeed

# Populate 'pagespeed.txt' file with URLs to query against API.
api="AIzaSyAUKgKypTLfllYA2Ggu62XUOAwQR6VvGjE"
url= "https://m.promovacances.com/"
device_mobile= "mobile"
device_desktop = "desktop"

#########################Pagespeed pour la hp######################################################
def wcv(url,data,device):
    #wb = load_workbook('suivi_perf_mep.xlsx')
    #ws = wb.active
    #wb.create_sheet('Data HP',0)
    #wb.create_sheet('suivi post mep',1)

    # This is the google pagespeed api url structure, using for loop to insert each url 
    # If no "strategy" parameter is included, the query by default returns desktop data.
    # requête sans l'api y = f'https://www.googleapis.com/pagespeedonline/v5/runPagespeed?url={line}&strategy=mobile'
    x = f'https://www.googleapis.com/pagespeedonline/v5/runPagespeed?url={url}&strategy={device}&locale=en&key={api}'
    print(f'Requesting {x}...')

    r = requests.get(x)

    final = r.json()
        
        
    urlid = final['id']
    split = urlid.split('?') # This splits the absolute url from the api key parameter
    urlid = split[0] # This reassigns urlid to the absolute url
    ID = f'URL ~ {urlid}'
    ID2 = str(urlid)


    #FCP
    urlfcp = final['lighthouseResult']['audits']['first-contentful-paint']['displayValue']
    FCP = f'First Contentful Paint ~ {str(urlfcp)}'
    FCP2 = str(urlfcp)
    #Suppression des caractères bloquant
    characters_fcp=" s"
    FCP2 =''.join( x for x in FCP2 if x not in characters_fcp)
    FCP2_2= float(FCP2)
    FCP2_virgule = FCP2_2 #.replace(".",",")

    #Time to Interactive
    urlfi = final['lighthouseResult']['audits']['interactive']['displayValue']
    FI = f'Timte To Interactive ~ {str(urlfi)}'
    FI2 = str(urlfi)
    #Suppression des caractères bloquant
    characters_fi="s"
    FI2 =''.join( x for x in FI2 if x not in characters_fi)
    FI2_2= float(FI2)
    FI2_virgule = FI2_2 #.replace(".",",")
    
     

    #LCP
    urllcp = final['lighthouseResult']['audits']['largest-contentful-paint']['displayValue']
    LCP = f'LCP ~ {str(urllcp)}'
    LCP2 = str(urllcp)
    characters_lcp=" s"
    LCP2 =''.join( x for x in LCP2 if x not in characters_lcp)
    LCP2= float(LCP2)
    LCP2_virgule = LCP2 #.replace(".",",")

    #Serveur reponse time
    urlsrt = final['lighthouseResult']['audits']['server-response-time']['displayValue']
    SRT = f'Server Reponse Time ~ {str(urlsrt)}'
    SRT2 = str(urlsrt)
    #Suppression des caractères bloquant
    characters_srt="Root document took ms"
    SRT2 =''.join( x for x in SRT2 if x not in characters_srt)
    #SRT2= float(SRT2)
    SRT2_virgule = SRT2 #.replace(".",",")

    #Speed Index
    urlsi = final['lighthouseResult']['audits']['speed-index']['displayValue']
    SI = f'Speed Index ~ {str(urlsi)}'
    SI2 = str(urlsi)
    #Suppression des caractères bloquant
    characters_si=" s"
    SI2 =''.join( x for x in SI2 if x not in characters_si)
    SI2= float(SI2)
    SI2_virgule = SI2 #.replace(".",",")

    #Cumulative Layout Shift
    urlcls = final['lighthouseResult']['audits']['cumulative-layout-shift']['displayValue']
    CLS = f'Cumulative Layout Shift ~ {str(urlcls)}'
    CLS2 = str(urlcls)
    CLS2= float(CLS2)
    CLS_virgule = CLS2 #.replace(".",",")

    #Total Blocking Time
    urltbt = final['lighthouseResult']['audits']['total-blocking-time']['displayValue']
    TBT = f'Total Blocking Time ~ {str(urltbt)}'
    TBT2 = str(urltbt)
    #Suppression charactères ms
    characters_tbt=" ms"
    TBT2 =''.join( x for x in TBT2 if x not in characters_tbt)
    #TBT2= float(TBT2)
    tbt2_virgule = TBT2.replace(",",".")
    tbt3_virgule= float(tbt2_virgule)


    #Lent, moyen, rapide FCP
    url_fcp_score = final["loadingExperience"]["metrics"]["FIRST_CONTENTFUL_PAINT_MS"]["category"]
    FCP_SCORE = f'FCP Score ~ {str(url_fcp_score)}'
    FCP_SCORE2 = str(url_fcp_score)




    #Lent, moyen, rapide LCP
    url_lcp_score = final["loadingExperience"]["metrics"]["LARGEST_CONTENTFUL_PAINT_MS"]["category"]
    LCP_SCORE = f'LCP Score ~ {str(url_lcp_score)}'
    LCP_SCORE2 = str(url_lcp_score)
    #Suppression des caractères bloquant



    #Lent, moyen, rapide CLS
    url_cls_score = final["loadingExperience"]["metrics"]["CUMULATIVE_LAYOUT_SHIFT_SCORE"]["category"]
    CLS_SCORE = f'CLS Score ~ {str(url_cls_score)}'
    CLS_SCORE2 = str(url_cls_score)
    #Suppression des caractères bloquant


    #Score global
    url_overall_score = final["lighthouseResult"]["categories"]["performance"]["score"] * 100
    OVERALL_SCORE = f'Score Global ~ {str(url_overall_score)}'
    OVERALL_SCORE2 = str(url_overall_score)
    #Suppression des caractères bloquant
    OVERALL_SCORE2= float(OVERALL_SCORE2)
    overall_score_virgule = OVERALL_SCORE2 #.replace(".",",")


    #Répertorie les requêtes réseau qui ont été faites lors du chargement de la page
    """listrequests = []
    for x in range (len(final["lighthouseResult"]["audits"]["network-requests"]["details"]["items"])):
        endtime = final["lighthouseResult"]["audits"]["network-requests"]["details"]["items"][x]["endTime"]
        starttime = final["lighthouseResult"]["audits"]["network-requests"]["details"]["items"][x]["startTime"]
        transfersize = final["lighthouseResult"]["audits"]["network-requests"]["details"]["items"][x]["transferSize"]
        resourcesize = final["lighthouseResult"]["audits"]["network-requests"]["details"]["items"][x]["resourceSize"]
        url = final["lighthouseResult"]["audits"]["network-requests"]["details"]["items"][x]["url"]
        list1 = [endtime, starttime, transfersize, resourcesize, url]
        listrequests.append(list1)
    """
    """#Total Bytes Weight ==> La boucle renverra une liste qui contient les différentes ressources avec leurs poids en octets.
    bytes_weight_score = final["lighthouseResult"]["audits"]["total-byte-weight"]["score"]
    bytes_weight = final["lighthouseResult"]["audits"]["total-byte-weight"]["displayValue"]

    listbytes = []
    for x in range (len(final["lighthouseResult"]["audits"]["total-byte-weight"]["details"]["items"])):
        url = final["lighthouseResult"]["audits"]["total-byte-weight"]["details"]["items"][x]["url"]
        bytes_total = final["lighthouseResult"]["audits"]["total-byte-weight"]["details"]["items"][x]["totalBytes"]
        list1 = [url, bytes_total]
        listbytes.append(list1)"""

    """ #Use of responsive images La boucle renverra une liste qui contient les différentes images avec leurs octets perdus et totaux.
    responsive_images_score = final["lighthouseResult"]["audits"]["uses-responsive-images"]["score"]
    responsive_image_savings = final["lighthouseResult"]["audits"]["uses-responsive-images"]["displayValue"]

    listresponsivesavings = []
    for x in range (len(final["lighthouseResult"]["audits"]["uses-responsive-images"]["details"]["items"])):
        url = final["lighthouseResult"]["audits"]["uses-responsive-images"]["details"]["items"][x]["url"]
        wastedbytes = final["lighthouseResult"]["audits"]["uses-responsive-images"]["details"]["items"][x]["wastedBytes"]
        totalbytes = final["lighthouseResult"]["audits"]["uses-responsive-images"]["details"]["items"][x]["totalBytes"]
        list1 = [url, wastedbytes, totalbytes]
        listresponsivesavings.append(list1)"""

    """#Render Blocking Resources La boucle renverra une liste qui contient les ressources qui bloquent le rendu avec leurs octets gaspillés et totaux.
    blocking_resources_score = final["lighthouseResult"]["audits"]["render-blocking-resources"]["score"]
    blocking_resoures_savings = final["lighthouseResult"]["audits"]["render-blocking-resources"]["displayValue"]

    listblockingresources = []
    for x in range (len(final["lighthouseResult"]["audits"]["render-blocking-resources"]["details"]["items"])):
        url = final["lighthouseResult"]["audits"]["render-blocking-resources"]["details"]["items"][x]["url"]
        totalbytes = final["lighthouseResult"]["audits"]["render-blocking-resources"]["details"]["items"][x]["totalBytes"]
        wastedbytes = final["lighthouseResult"]["audits"]["render-blocking-resources"]["details"]["items"][x]["wastedMs"]
        list1 = [url, totalbytes, wastedbytes]
        listblockingresources.append(list1)


    #Use of Rel Preload La boucle renverra une liste avec les ressources qui peuvent être préchargées et leurs millisecondes perdues.
    rel_preload_score = final["lighthouseResult"]["audits"]["uses-rel-preload"]["score"]
    rel_preload_savings = final["lighthouseResult"]["audits"]["uses-rel-preload"]["displayValue"]

    listrelpreload = []
    for x in range (len(final["lighthouseResult"]["audits"]["uses-rel-preload"]["details"]["items"])):
        url = final["lighthouseResult"]["audits"]["uses-rel-preload"]["details"]["items"][x]["url"]
        wastedms = final["lighthouseResult"]["audits"]["uses-rel-preload"]["details"]["items"][x]["wastedMs"]
        list1 = [url, wastedms]
        listrelpreload.append(list1)


    #Redirects La boucle renverra une liste qui contient les redirections et les millisecondes perdues sur chaque redirection.
    redirects_score = final["lighthouseResult"]["audits"]["redirects"]["score"]
    redirect_savings = final["lighthouseResult"]["audits"]["redirects"]["displayValue"]

    listredirects = []
    for x in range (len(final["lighthouseResult"]["audits"]["redirects"]["details"]["items"])):
        url = final["lighthouseResult"]["audits"]["redirects"]["details"]["items"][x]["url"]
        wastedms = final["lighthouseResult"]["audits"]["redirects"]["details"]["items"][x]["wastedMs"]
        list1 = [url,wastedms]
        listredirects.append(list1)
    """

    """#Javascript inutilisé La boucle renverra une liste avec les fichiers Javascript inutilisés, leur nombre total d'octets, les octets perdus et le pourcentage d'octets perdus.
    unused_js_score = final["lighthouseResult"]["audits"]["unused-javascript"]["score"]
    unused_js_savings = final["lighthouseResult"]["audits"]["unused-javascript"]["displayValue"]

    listunusedjavascript = []
    for x in range (len(final["lighthouseResult"]["audits"]["unused-javascript"]["details"]["items"])):
        url = final["lighthouseResult"]["audits"]["unused-javascript"]["details"]["items"][x]["url"]
        totalbytes = final["lighthouseResult"]["audits"]["unused-javascript"]["details"]["items"][x]["totalBytes"]
        wastedbytes = final["lighthouseResult"]["audits"]["unused-javascript"]["details"]["items"][x]["wastedBytes"]
        wastedpercentage= final["lighthouseResult"]["audits"]["unused-javascript"]["details"]["items"][x]["wastedPercent"]
        list1 = [url, totalbytes, wastedbytes, wastedpercentage]
        listunusedjavascript.append(list1)"""


    """#Liste des javascript pouvant être minifié, La boucle renverra une liste qui contient les ressources Javascript qui peuvent être minifiées, le nombre total d'octets, le nombre d'octets perdus et le pourcentage d'octets perdus
    unminified_javascript_score = final["lighthouseResult"]["audits"]["unminified-javascript"]["score"]
    unminified_javascript_savings = final["lighthouseResult"]["audits"]["unminified-javascript"]["displayValue"]

    listunminifiedjavascript = []
    for x in range (len(final["lighthouseResult"]["audits"]["unminified-javascript"]["details"]["items"])):
        url = final["lighthouseResult"]["audits"]["unminified-javascript"]["details"]["items"][x]["url"]
        wastedbytes = final["lighthouseResult"]["audits"]["unminified-javascript"]["details"]["items"][x]["wastedBytes"]
        totalbytes = final["lighthouseResult"]["audits"]["unminified-javascript"]["details"]["items"][x]["totalBytes"]
        wastedpercent = totalbytes = final["lighthouseResult"]["audits"]["unminified-javascript"]["details"]["items"][x]["wastedPercent"]
        list1 = [url, wastedbytes, totalbytes, wastedpercent]
        listunminifiedjavascript.append(list1)"""


    #Long Tasks La boucle retournera une liste qui contient les tâches longues et leurs durées.
#    long_tasks = final["lighthouseResult"]["audits"]["long-tasks"]["displayValue"]

    listlongtasks = []
    for x in range (len(final["lighthouseResult"]["audits"]["long-tasks"]["details"]["items"])):
        url = final["lighthouseResult"]["audits"]["long-tasks"]["details"]["items"][x]["url"]
        duration = final["lighthouseResult"]["audits"]["long-tasks"]["details"]["items"][x]["duration"]
        list1 = [url, duration]
        listlongtasks.append(list1)
        
            
    


    print(ID2) 
    print('________________________________________________________________________________________________________________')
    print("FCP")
    print('________________________________________________________________________________________________________________')
    print(FCP2)
    print('________________________________________________________________________________________________________________')
    print('FI Score')
    print('________________________________________________________________________________________________________________')
    print(FI2)
    print('________________________________________________________________________________________________________________')
    print('LCP Score')
    print('________________________________________________________________________________________________________________')
    print(LCP2)
    print('________________________________________________________________________________________________________________')
    print('SRT Score')
    print('________________________________________________________________________________________________________________')
    print(SRT2)
    print('________________________________________________________________________________________________________________')
    print('SI Score')
    print('________________________________________________________________________________________________________________')
    print(SI2)
    print('________________________________________________________________________________________________________________')
    print('CLS Score')
    print('________________________________________________________________________________________________________________')
    print(CLS2)
    print('________________________________________________________________________________________________________________')
    print('TBT Score')
    print('________________________________________________________________________________________________________________')
    print(TBT2)
    print('________________________________________________________________________________________________________________')
    print('FCP Score')
    print('________________________________________________________________________________________________________________')
    print(FCP_SCORE2)
    print('________________________________________________________________________________________________________________')
    print('LCP Score')
    print('________________________________________________________________________________________________________________')
    print(LCP_SCORE2)
    print('________________________________________________________________________________________________________________')
    print('Cls Score')
    print('________________________________________________________________________________________________________________')
    print(CLS_SCORE2)
    print('________________________________________________________________________________________________________________')
    print("Score global")
    print('________________________________________________________________________________________________________________')
    print(OVERALL_SCORE2)
    print('________________________________________________________________________________________________________________')
    print("Liste des requêtes réseau qui ont été faites lors du chargement de la page")
    print('________________________________________________________________________________________________________________')
    #print(listrequests)
    print('________________________________________________________________________________________________________________')
    #print('liste qui contient les différentes ressources avec leurs poids en octets.')
    #print(listbytes)
    #print(listresponsivesavings)
    print("liste qui contient les ressources qui bloquent le rendu avec leurs octets gaspillés et totaux")
    #print(listblockingresources)
    print('________________________________________________________________________________________________________________')
    #print("les ressources qui peuvent être préchargées et leurs millisecondes perdues")
    #print(listrelpreload)
    #print("liste des ressources en redirections")
    #print(listredirects)
    #print("Javascript inutilisé: liste avec les fichiers Javascript inutilisés")
    #print(listunusedjavascript)
    #print("Liste des javascript pouvant être minifié")
    #print(listunminifiedjavascript)
    print('________________________________________________________________________________________________________________')
    print("Tâches longues")
    print(listlongtasks)
    print('________________________________________________________________________________________________________________')




    time.sleep(15)
    date = datetime.today().strftime('%Y-%m-%d %H:%M:%S')
    #data = wb['Data HP']
    data['A1'] = "URL"
    data['A2'] = ID2
    data['B1'] = "Date"
    data['B2'] = date
    data['C1'] = "Time to Interactive"
    data['C2'] = FI2_virgule
    data['D1'] = "Server Reponse Time"
    data['D2'] = SRT2_virgule
    data['E1'] = "Speed Index"
    data['E2'] = SI2_virgule
    data['F1'] = "Total Blocking Time"
    data['F2'] = tbt3_virgule
    data['G1'] = "FCP"
    data['G2'] = FCP2_virgule
    data['H1'] = "LCP"
    data['H2'] = LCP2_virgule
    data['I1'] = "CLS"
    data['I2'] = CLS_virgule
    data['J1'] = "Score FCP"
    data['J2'] = FCP_SCORE2
    data['K1'] = "LCP Score"
    data['K2'] = LCP_SCORE2
    data['L1'] = "CLS Score"
    data['L2'] = CLS_SCORE2
    data['M1'] = "Score Global"
    data['M2'] = overall_score_virgule



   

