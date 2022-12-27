import os

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill

from google_sheet import clean_google_sheet, excel_to_googlesheet
from pagespeed_api_hp import wcv

whiteFill = PatternFill(fgColor="0070C0", bgColor="0070C0", fill_type="solid")
orangeFill = PatternFill(fgColor="0064DC", bgColor="0064DC", fill_type="solid")
greyFill = PatternFill(fgColor="D9D9D9", bgColor="D9D9D9", fill_type="solid")
grey_fill_light= PatternFill(fgColor="F2F2F2", bgColor="F2F2F2", fill_type="solid")
text_color = Font(color="FFFFFF")
text_color_red = Font(color="FF0000")
yellowFill = PatternFill(fgColor="FFFF00", bgColor="FFFF00", fill_type="solid")    
yellow_fill_light = PatternFill(fgColor="FFFF99", bgColor="FFFF99", fill_type="solid")
grey_date = PatternFill(fgColor="D9D9D9", bgColor="D9D9D9", fill_type="solid")
text_bold = Font(bold=True)# mettre la police en gras

device_mobile= "mobile"
device_desktop = "desktop"
name_sheet = "Historisation_temps_de_chargement_Fram"
# declaration
#fichiers
chem_result = 'result.xlsx'
chem_suivi_perf_mep = 'suivi_perf_mep.xlsx'
# anglets Mobile
anglet_HP_mobile = 'Data HP_mobile'
anglet_SL_mobile = 'Data SL_mobile'
anglet_FP_mobile = 'Data FP_mobile'
anglet_LD_mobile = 'Data LD_mobile'
anglet_HPT_mobile = 'Data HPT_mobile'
anglet_QDP_mobile = 'Data QDP_mobile'

# anglets Desktop
anglet_HP = 'Data HP'
anglet_SL = 'Data SL'
anglet_FP = 'Data FP'
anglet_LD = 'Data LD'
anglet_HPT = 'Data HPT'
anglet_QDP = 'Data QDP'

# Color mobile
color_hp_mobile ="00B050"
color_sl_mobile ="00B060"
color_fp_mobile ="00B070"
color_ld_mobile ="00B080"
color_hpt_mobile ="00B090"
color_qdp_mobile ="00B100"

# Color desktop
color_hp ="00B050"
color_sl ="00B060"
color_fp ="00B070"
color_ld ="00B080"
color_hpt ="00B090"
color_qdp ="00B100"


# Documentation: https://developers.google.com/speed/docs/insights/v5/get-started

# JSON paths: https://developers.google.com/speed/docs/insights/v4/reference/pagespeedapi/runpagespeed

# Populate 'pagespeed.txt' file with URLs to query against API.
api="AIzaSyAUKgKypTLfllYA2Ggu62XUOAwQR6VvGjE"

#Création du fichier Excel 
wb  = Workbook()
#Mobile
wb.create_sheet(anglet_HP_mobile,0)
wb.create_sheet(anglet_SL_mobile,1)
wb.create_sheet(anglet_FP_mobile,2)
wb.create_sheet(anglet_LD_mobile,3)
wb.create_sheet(anglet_HPT_mobile,4)
wb.create_sheet(anglet_QDP_mobile,5)

#Desktop
wb.create_sheet(anglet_HP,0)
wb.create_sheet(anglet_SL,1)
wb.create_sheet(anglet_FP,2)
wb.create_sheet(anglet_LD,3)
wb.create_sheet(anglet_HPT,4)
wb.create_sheet(anglet_QDP,5)

#Configuration de la partie Mobile ____________________________________________________________________________________________________________________________
onglet_data_hp_mobile = wb[anglet_HP_mobile]
#Coloration de l'onglet
onglet_data_hp_mobile.sheet_properties.tabColor =color_hp_mobile
#Insertion des données pagespeedinsight
wcv("https://m.fram.fr/",onglet_data_hp_mobile,device_mobile)


#Insertion des données pagespeedinsight
onglet_data_sl_mobile = wb[anglet_SL_mobile]
#Coloration de l'onglet 
onglet_data_sl_mobile.sheet_properties.tabColor =color_sl_mobile
#Insertion des données pagespeedinsight
wcv("https://m.fram.fr/republique-dominicaine/sejour-vacances-voyage-republique-dominicaine-pas-cher-2/",onglet_data_sl_mobile,device_mobile)

#Insertion des données pagespeedinsight
onglet_data_fp_mobile = wb[anglet_FP_mobile]
#Coloration de l'onglet 
onglet_data_fp_mobile.sheet_properties.tabColor =color_fp_mobile
#Insertion des données pagespeedinsight
wcv("https://m.fram.fr/hotel-eden-star-tunisie-zarzis-56952.html",onglet_data_fp_mobile,device_mobile)

#Insertion des données pagespeedinsight
onglet_data_ld_mobile = wb[anglet_LD_mobile]
#Coloration de l'onglet 
onglet_data_ld_mobile.sheet_properties.tabColor =color_ld_mobile
#Insertion des données pagespeedinsight
wcv("https://m.fram.fr/sejour-vacances-voyage-derniere-minute/",onglet_data_ld_mobile,device_mobile)


#Insertion des données pagespeedinsight
onglet_data_hpt_mobile = wb[anglet_HPT_mobile]
#Coloration de l'onglet 
onglet_data_hpt_mobile.sheet_properties.tabColor =color_hpt_mobile
#Insertion des données pagespeedinsight
wcv("https://m.fram.fr/sejour/vacances-voyage-all-inclusive-tout-compris/",onglet_data_hpt_mobile,device_mobile)


#Insertion des données pagespeedinsight
onglet_data_qdp_mobile = wb[anglet_QDP_mobile]
#Coloration de l'onglet 
onglet_data_qdp_mobile.sheet_properties.tabColor =color_qdp_mobile
#Insertion des données pagespeedinsight
wcv("https://m.fram.fr/ou-quand-partir-en-voyage/vacances-juillet/",onglet_data_qdp_mobile,device_mobile)



#Configuration de la partie Desktop  ____________________________________________________________________________________________________________________________
onglet_data_hp= wb[anglet_HP]
#Coloration de l'onglet
onglet_data_hp.sheet_properties.tabColor =color_hp
#Insertion des données pagespeedinsight
wcv("https://www.fram.fr/",onglet_data_hp,device_desktop)


#Insertion des données pagespeedinsight
onglet_data_sl = wb[anglet_SL]
#Coloration de l'onglet 
onglet_data_sl.sheet_properties.tabColor =color_sl
#Insertion des données pagespeedinsight
wcv("https://www.fram.fr/republique-dominicaine/sejour-vacances-voyage-republique-dominicaine-pas-cher-2/",onglet_data_sl,device_desktop)

#Insertion des données pagespeedinsight
onglet_data_fp = wb[anglet_FP]
#Coloration de l'onglet 
onglet_data_fp.sheet_properties.tabColor =color_fp
#Insertion des données pagespeedinsight
wcv("https://www.fram.fr/hotel-eden-star-tunisie-zarzis-56952.html",onglet_data_fp,device_desktop)

#Insertion des données pagespeedinsight
onglet_data_ld= wb[anglet_LD]
#Coloration de l'onglet 
onglet_data_ld.sheet_properties.tabColor =color_ld
#Insertion des données pagespeedinsight
wcv("https://www.fram.fr/sejour-vacances-voyage-derniere-minute/",onglet_data_ld,device_desktop)


#Insertion des données pagespeedinsight
onglet_data_hpt= wb[anglet_HPT]
#Coloration de l'onglet 
onglet_data_hpt.sheet_properties.tabColor =color_hpt
#Insertion des données pagespeedinsight
wcv("https://www.fram.fr/sejour/vacances-voyage-all-inclusive-tout-compris/",onglet_data_hpt,device_desktop)


#Insertion des données pagespeedinsight
onglet_data_qdp= wb[anglet_QDP]
#Coloration de l'onglet 
onglet_data_qdp.sheet_properties.tabColor =color_qdp
#Insertion des données pagespeedinsight
wcv("https://www.fram.fr/ou-quand-partir-en-voyage/vacances-juillet/",onglet_data_qdp,device_desktop)




#Sauvegarde du fichier result.xlsx___________________________________________________________________________________________________________________________________
wb.save(chem_result)

if os.path.isfile(chem_suivi_perf_mep) and os.access(chem_suivi_perf_mep, os.R_OK):
    print("Lecture du fichier:",chem_suivi_perf_mep)
    wb_hist = load_workbook(chem_suivi_perf_mep)
    #Coloration de l'onglet mobile
    onglet_data_hist_hp_mobile = wb_hist[anglet_HP_mobile]   
    onglet_data_hist_sl_mobile = wb_hist[anglet_SL_mobile]
    onglet_data_hist_fp_mobile = wb_hist[anglet_FP_mobile]

    onglet_data_hist_ld_mobile = wb_hist[anglet_LD_mobile]
    onglet_data_hist_hpt_mobile = wb_hist[anglet_HPT_mobile]
    onglet_data_hist_qdp_mobile = wb_hist[anglet_QDP_mobile]

    onglet_data_hist_hp_mobile.sheet_properties.tabColor =color_hp    
    onglet_data_hist_sl_mobile.sheet_properties.tabColor =color_sl
    onglet_data_hist_fp_mobile.sheet_properties.tabColor =color_fp

    onglet_data_hist_ld_mobile.sheet_properties.tabColor =color_ld    
    onglet_data_hist_hpt_mobile.sheet_properties.tabColor =color_hpt
    onglet_data_hist_qdp_mobile.sheet_properties.tabColor =color_qdp
    
   #Coloration de l'onglet desktop
    onglet_data_hist_hp = wb_hist[anglet_HP]   
    onglet_data_hist_sl = wb_hist[anglet_SL]
    onglet_data_hist_fp = wb_hist[anglet_FP]

    onglet_data_hist_ld = wb_hist[anglet_LD]
    onglet_data_hist_hpt = wb_hist[anglet_HPT]
    onglet_data_hist_qdp = wb_hist[anglet_QDP]

    onglet_data_hist_hp.sheet_properties.tabColor =color_hp    
    onglet_data_hist_sl.sheet_properties.tabColor =color_sl
    onglet_data_hist_fp.sheet_properties.tabColor =color_fp

    onglet_data_hist_ld.sheet_properties.tabColor =color_ld    
    onglet_data_hist_hpt.sheet_properties.tabColor =color_hpt
    onglet_data_hist_qdp.sheet_properties.tabColor =color_qdp


    wb_hist.save(chem_suivi_perf_mep)
else:
    print("Creation du fichier:",chem_suivi_perf_mep)
    wb_hist = Workbook()
    #Onglet mobile
    wb_hist.create_sheet(anglet_HP_mobile,0)
    #Onglet Desktop
    wb_hist.create_sheet(anglet_HP,1)
    #Onglet mobile
    wb_hist.create_sheet(anglet_SL_mobile,2)
    #Onglet Desktop
    wb_hist.create_sheet(anglet_SL,3)
    #Onglet mobile
    wb_hist.create_sheet(anglet_FP_mobile,4)
    #Onglet Desktop
    wb_hist.create_sheet(anglet_FP,5)
    #Onglet mobile
    wb_hist.create_sheet(anglet_LD_mobile,6)
    #Onglet Desktop
    wb_hist.create_sheet(anglet_LD,7)
    #Onglet mobile
    wb_hist.create_sheet(anglet_HPT_mobile,8)
    #Onglet Desktop
    wb_hist.create_sheet(anglet_HPT,9)
    #Onglet mobile
    wb_hist.create_sheet(anglet_QDP_mobile,10)
    #Onglet Desktop
    wb_hist.create_sheet(anglet_QDP,11)

    #Coloration de l'onglet Mobile
    onglet_data_hist_hp_mobile = wb_hist[anglet_HP_mobile]   
    onglet_data_hist_sl_mobile = wb_hist[anglet_SL_mobile]
    onglet_data_hist_fp_mobile = wb_hist[anglet_FP_mobile]
    onglet_data_hist_ld_mobile = wb_hist[anglet_LD_mobile]   
    onglet_data_hist_hpt_mobile = wb_hist[anglet_HPT_mobile]
    onglet_data_hist_qdp_mobile = wb_hist[anglet_QDP_mobile]

    onglet_data_hist_hp_mobile.sheet_properties.tabColor =color_hp_mobile   
    onglet_data_hist_sl_mobile.sheet_properties.tabColor =color_sl_mobile
    onglet_data_hist_fp_mobile.sheet_properties.tabColor =color_fp_mobile
    onglet_data_hist_ld_mobile.sheet_properties.tabColor =color_ld_mobile    
    onglet_data_hist_hpt_mobile.sheet_properties.tabColor =color_hpt_mobile
    onglet_data_hist_qdp_mobile.sheet_properties.tabColor =color_qdp_mobile

    #Coloration de l'onglet Desktop
    onglet_data_hist_hp = wb_hist[anglet_HP]   
    onglet_data_hist_sl = wb_hist[anglet_SL]
    onglet_data_hist_fp = wb_hist[anglet_FP]
    onglet_data_hist_ld = wb_hist[anglet_LD]   
    onglet_data_hist_hpt = wb_hist[anglet_HPT]
    onglet_data_hist_qdp = wb_hist[anglet_QDP]

    onglet_data_hist_hp.sheet_properties.tabColor =color_hp    
    onglet_data_hist_sl.sheet_properties.tabColor =color_sl
    onglet_data_hist_fp.sheet_properties.tabColor =color_fp
    onglet_data_hist_ld.sheet_properties.tabColor =color_ld    
    onglet_data_hist_hpt.sheet_properties.tabColor =color_hpt
    onglet_data_hist_qdp.sheet_properties.tabColor =color_qdp

    wb_hist.save(chem_suivi_perf_mep)


# pour pour recuperer le dataframe du  fichier result Mobile
df_result_HP_mobile = pd.read_excel(chem_result,anglet_HP_mobile,engine='openpyxl')
df_result_SL_mobile = pd.read_excel(chem_result,anglet_SL_mobile,engine='openpyxl')
df_result_FP_mobile = pd.read_excel(chem_result,anglet_FP_mobile,engine='openpyxl')
df_result_LD_mobile = pd.read_excel(chem_result,anglet_LD_mobile,engine='openpyxl')
df_result_HPT_mobile = pd.read_excel(chem_result,anglet_HPT_mobile,engine='openpyxl')
df_result_QDP_mobile = pd.read_excel(chem_result,anglet_QDP_mobile,engine='openpyxl')

# pour pour recuperer le dataframe du  fichier result Mobile
df_result_HP = pd.read_excel(chem_result,anglet_HP,engine='openpyxl')
df_result_SL = pd.read_excel(chem_result,anglet_SL,engine='openpyxl')
df_result_FP = pd.read_excel(chem_result,anglet_FP,engine='openpyxl')
df_result_LD = pd.read_excel(chem_result,anglet_LD,engine='openpyxl')
df_result_HPT = pd.read_excel(chem_result,anglet_HPT,engine='openpyxl')
df_result_QDP = pd.read_excel(chem_result,anglet_QDP,engine='openpyxl')



# pour pour recuperer le dataframe du  fichier result Mobile
df_result_HP_mobile = pd.read_excel(chem_result,anglet_HP_mobile,engine='openpyxl')
df_result_SL_mobile = pd.read_excel(chem_result,anglet_SL_mobile,engine='openpyxl')
df_result_FP_mobile = pd.read_excel(chem_result,anglet_FP_mobile,engine='openpyxl')
df_result_LD_mobile = pd.read_excel(chem_result,anglet_LD_mobile,engine='openpyxl')
df_result_HPT_mobile = pd.read_excel(chem_result,anglet_HPT_mobile,engine='openpyxl')
df_result_QDP_mobile = pd.read_excel(chem_result,anglet_QDP_mobile,engine='openpyxl')

# pour pour recuperer le dataframe du  fichier result Desktop
df_result_HP = pd.read_excel(chem_result,anglet_HP,engine='openpyxl')
df_result_SL = pd.read_excel(chem_result,anglet_SL,engine='openpyxl')
df_result_FP = pd.read_excel(chem_result,anglet_FP,engine='openpyxl')
df_result_LD = pd.read_excel(chem_result,anglet_LD,engine='openpyxl')
df_result_HPT = pd.read_excel(chem_result,anglet_HPT,engine='openpyxl')
df_result_QDP = pd.read_excel(chem_result,anglet_QDP,engine='openpyxl')



# pour recuperer le dataframe du  fichier suivi_perf_mep Mobile
df_mep_3_hist_Data_HP_mobile = pd.read_excel(chem_suivi_perf_mep,anglet_HP_mobile,engine='openpyxl')
df_mep_3_hist_Data_SL_mobile = pd.read_excel(chem_suivi_perf_mep,anglet_SL_mobile,engine='openpyxl')
df_mep_3_hist_Data_FP_mobile = pd.read_excel(chem_suivi_perf_mep,anglet_FP_mobile,engine='openpyxl')
df_mep_3_hist_Data_LD_mobile = pd.read_excel(chem_suivi_perf_mep,anglet_LD_mobile,engine='openpyxl')
df_mep_3_hist_Data_HPT_mobile = pd.read_excel(chem_suivi_perf_mep,anglet_HPT_mobile,engine='openpyxl')
df_mep_3_hist_Data_QDP_mobile = pd.read_excel(chem_suivi_perf_mep,anglet_QDP_mobile,engine='openpyxl')

# pour recuperer le dataframe du  fichier suivi_perf_mep Desktop
df_mep_3_hist_Data_HP = pd.read_excel(chem_suivi_perf_mep,anglet_HP,engine='openpyxl')
df_mep_3_hist_Data_SL = pd.read_excel(chem_suivi_perf_mep,anglet_SL,engine='openpyxl')
df_mep_3_hist_Data_FP = pd.read_excel(chem_suivi_perf_mep,anglet_FP,engine='openpyxl')
df_mep_3_hist_Data_LD = pd.read_excel(chem_suivi_perf_mep,anglet_LD,engine='openpyxl')
df_mep_3_hist_Data_HPT = pd.read_excel(chem_suivi_perf_mep,anglet_HPT,engine='openpyxl')
df_mep_3_hist_Data_QDP = pd.read_excel(chem_suivi_perf_mep,anglet_QDP,engine='openpyxl')


# concaténation des deux dataframe Mobile
df_mep_3_hist_Data_HP_mobile = pd.concat([df_mep_3_hist_Data_HP_mobile,df_result_HP_mobile],ignore_index=True)
df_mep_3_hist_Data_SL_mobile = pd.concat([df_mep_3_hist_Data_SL_mobile,df_result_SL_mobile],ignore_index=True)
df_mep_3_hist_Data_FP_mobile = pd.concat([df_mep_3_hist_Data_FP_mobile,df_result_FP_mobile],ignore_index=True)
df_mep_3_hist_Data_LD_mobile = pd.concat([df_mep_3_hist_Data_LD_mobile,df_result_LD_mobile],ignore_index=True)
df_mep_3_hist_Data_HPT_mobile = pd.concat([df_mep_3_hist_Data_HPT_mobile,df_result_HPT_mobile],ignore_index=True)
df_mep_3_hist_Data_QDP_mobile = pd.concat([df_mep_3_hist_Data_QDP_mobile,df_result_QDP_mobile],ignore_index=True)

# concaténation des deux dataframe Desktop
df_mep_3_hist_Data_HP = pd.concat([df_mep_3_hist_Data_HP,df_result_HP],ignore_index=True)
df_mep_3_hist_Data_SL = pd.concat([df_mep_3_hist_Data_SL,df_result_SL],ignore_index=True)
df_mep_3_hist_Data_FP = pd.concat([df_mep_3_hist_Data_FP,df_result_FP],ignore_index=True)
df_mep_3_hist_Data_LD = pd.concat([df_mep_3_hist_Data_LD,df_result_LD],ignore_index=True)
df_mep_3_hist_Data_HPT = pd.concat([df_mep_3_hist_Data_HPT,df_result_HPT],ignore_index=True)
df_mep_3_hist_Data_QDP = pd.concat([df_mep_3_hist_Data_QDP,df_result_QDP],ignore_index=True)


# sauvegarder les data fram dan le fichier suivi_perf_mep
with pd.ExcelWriter(chem_suivi_perf_mep) as writer: 
    df_mep_3_hist_Data_HP_mobile.to_excel(writer,sheet_name=anglet_HP_mobile, index=False, engine='xlsxwriter')
    df_mep_3_hist_Data_SL_mobile.to_excel(writer,sheet_name=anglet_SL_mobile, index=False, engine='xlsxwriter')
    df_mep_3_hist_Data_FP_mobile.to_excel(writer,sheet_name=anglet_FP_mobile, index=False, engine='xlsxwriter')
    df_mep_3_hist_Data_LD_mobile.to_excel(writer,sheet_name=anglet_LD_mobile, index=False, engine='xlsxwriter')
    df_mep_3_hist_Data_HPT_mobile.to_excel(writer,sheet_name=anglet_HPT_mobile, index=False, engine='xlsxwriter')
    df_mep_3_hist_Data_QDP_mobile.to_excel(writer,sheet_name=anglet_QDP_mobile, index=False, engine='xlsxwriter')

    df_mep_3_hist_Data_HP.to_excel(writer,sheet_name=anglet_HP, index=False, engine='xlsxwriter')
    df_mep_3_hist_Data_SL.to_excel(writer,sheet_name=anglet_SL, index=False, engine='xlsxwriter')
    df_mep_3_hist_Data_FP.to_excel(writer,sheet_name=anglet_FP, index=False, engine='xlsxwriter')
    df_mep_3_hist_Data_LD.to_excel(writer,sheet_name=anglet_LD, index=False, engine='xlsxwriter')
    df_mep_3_hist_Data_HPT.to_excel(writer,sheet_name=anglet_HPT, index=False, engine='xlsxwriter')
    df_mep_3_hist_Data_QDP.to_excel(writer,sheet_name=anglet_QDP, index=False, engine='xlsxwriter')


    
    
#Coloration de l'onglet 
wb_hist = load_workbook(chem_suivi_perf_mep)
#Mobile
onglet_data_hist_hp_mobile = wb_hist[anglet_HP_mobile]   
onglet_data_hist_sl_mobile = wb_hist[anglet_SL_mobile]
onglet_data_hist_fp_mobile = wb_hist[anglet_FP_mobile]
onglet_data_hist_ld_mobile = wb_hist[anglet_LD_mobile]   
onglet_data_hist_hpt_mobile = wb_hist[anglet_HPT_mobile]
onglet_data_hist_qdp_mobile = wb_hist[anglet_QDP_mobile]
#Desktop
onglet_data_hist_hp = wb_hist[anglet_HP]   
onglet_data_hist_sl = wb_hist[anglet_SL]
onglet_data_hist_fp = wb_hist[anglet_FP]
onglet_data_hist_ld = wb_hist[anglet_LD]   
onglet_data_hist_hpt = wb_hist[anglet_HPT]
onglet_data_hist_qdp = wb_hist[anglet_QDP]

#Mobile
onglet_data_hist_hp_mobile.sheet_properties.tabColor =color_hp_mobile    
onglet_data_hist_sl_mobile.sheet_properties.tabColor =color_sl_mobile
onglet_data_hist_fp_mobile.sheet_properties.tabColor =color_fp_mobile
onglet_data_hist_ld_mobile.sheet_properties.tabColor =color_ld_mobile    
onglet_data_hist_hpt_mobile.sheet_properties.tabColor =color_hpt_mobile
onglet_data_hist_qdp_mobile.sheet_properties.tabColor =color_qdp_mobile

#Desktop
onglet_data_hist_hp.sheet_properties.tabColor =color_hp    
onglet_data_hist_sl.sheet_properties.tabColor =color_sl
onglet_data_hist_fp.sheet_properties.tabColor =color_fp
onglet_data_hist_ld.sheet_properties.tabColor =color_ld    
onglet_data_hist_hpt.sheet_properties.tabColor =color_hpt
onglet_data_hist_qdp.sheet_properties.tabColor =color_qdp

wb_hist.save(chem_suivi_perf_mep)


#Open an xlsx for reading
wb = load_workbook('suivi_perf_mep.xlsx')
#Get the current Active Sheet
data_hp_s = wb['Data HP']
data_sl_s = wb['Data SL']
data_fp_s = wb['Data FP']
data_ld_s = wb['Data LD']
data_hpt_s = wb['Data HPT']
data_qdp_s = wb['Data QDP']
data_hp_mobile_s = wb['Data HP_mobile']
data_sl_mobile_s = wb['Data SL_mobile']
data_fp_mobile_s = wb['Data FP_mobile']
data_ld_mobile_s = wb['Data LD_mobile']
data_hpt_mobile_s = wb['Data HPT_mobile']
data_qdp_mobile_s = wb['Data QDP_mobile']


def entete(onglet):
    #Suppression du quadrillage
    onglet.sheet_view.showGridLines = False
    #Couleur de fond : 
    onglet['A1'].fill = orangeFill
    onglet['B1'].fill = orangeFill
    onglet['C1'].fill = orangeFill
    onglet['D1'].fill = orangeFill
    onglet['E1'].fill = orangeFill
    onglet['F1'].fill = orangeFill
    onglet['G1'].fill = orangeFill
    onglet['H1'].fill = orangeFill
    onglet['I1'].fill = orangeFill
    onglet['J1'].fill = orangeFill
    onglet['K1'].fill = orangeFill
    onglet['L1'].fill = orangeFill
    onglet['M1'].fill = orangeFill


    #Couleur texte
    onglet['A1'].font = text_color
    onglet['B1'].font = text_color
    onglet['C1'].font = text_color
    onglet['D1'].font = text_color
    onglet['E1'].font = text_color
    onglet['F1'].font = text_color
    onglet['G1'].font = text_color
    onglet['H1'].font = text_color
    onglet['I1'].font = text_color
    onglet['J1'].font = text_color
    onglet['K1'].font = text_color
    onglet['L1'].font = text_color
    onglet['M1'].font = text_color
    #Dimensionnement colonnes 
    onglet.column_dimensions["A"].width=29
    onglet.column_dimensions["B"].width=17.43
    onglet.column_dimensions["C"].width=17.14
    onglet.column_dimensions["D"].width=24
    onglet.column_dimensions["E"].width=11.29
    onglet.column_dimensions["F"].width=17.57
    onglet.column_dimensions["G"].width=4.14
    onglet.column_dimensions["H"].width=4.14
    onglet.column_dimensions["I"].width=3.29
    onglet.column_dimensions["J"].width=8.86
    onglet.column_dimensions["K"].width=8.71
    onglet.column_dimensions["L"].width=8.57
    onglet.column_dimensions["M"].width=11.43
    for r in range(1,42):
        onglet[f'C{r}'].number_format ='"$"#,##0_);[Red]("$"#,##0)'

entete(data_hp_s)
entete(data_sl_s)
entete(data_fp_s)
entete(data_ld_s)
entete(data_hpt_s)
entete(data_qdp_s)
entete(data_hp_mobile_s)
entete(data_sl_mobile_s)
entete(data_fp_mobile_s)
entete(data_ld_mobile_s)
entete(data_hpt_mobile_s)
entete(data_qdp_mobile_s)

wb.save('suivi_perf_mep.xlsx')


clean_google_sheet(name_sheet)
excel_to_googlesheet(name_sheet)













